"""Parses a CSV/XLSX file back into Full-View-shaped rows for the native
Import dialog (User-Wunsch, 2026-09-05: "Kann man hier auch ein Import
Button einfügen mit Vorschau?"). Deliberately reads the SAME shape
full_view_export.py's build_full_view_csv/build_full_view_xlsx already
produce (character rows x "[Type] Title" activity columns, "1"/"0"/""
cells) -- a round trip, not a new format, so a file exported from Full
View, edited in Excel/Sheets, and reimported here just works.

No live write-back from the exported browser page itself is possible (a
static file:// page has no channel into this running process, see
full_view_export.py's own module docstring) -- Sync happens natively here
instead, via MainWindow.sync_full_view_import, which is the only thing
that can actually touch the real profile.
"""

import re
from dataclasses import dataclass, field

from .full_view_export import _T as _EXPORT_T

_TYPE_KEYS = ("daily", "weekly", "season")
_HEADER_RE = re.compile(r"^\[(\w+)\]\s*(.+)$")


def _all_unassigned_labels() -> set[str]:
    """Every language's "unassigned" display text (User could reimport a
    file exported in a different language than the app's current one, or
    just never touch that cell) -- matched case-insensitively."""
    return {texts["unassigned"].strip().lower() for texts in _EXPORT_T.values()}


@dataclass
class ImportEntry:
    character: str | None  # None = unassigned
    schedule: str
    title: str
    done: bool


@dataclass
class ParsedImport:
    entries: list[ImportEntry] = field(default_factory=list)
    unmatched_characters: list[str] = field(default_factory=list)  # rows skipped, name matched nothing real


def _parse_header_cell(cell: str) -> tuple[str, str] | None:
    match = _HEADER_RE.match((cell or "").strip())
    if not match:
        return None
    type_key = match.group(1).strip().lower()
    if type_key not in _TYPE_KEYS:
        type_key = "daily"
    title = match.group(2).strip()
    if not title:
        return None
    return type_key, title


def _resolve_character(raw: str, characters: list[str], unassigned_labels: set[str]) -> tuple[str | None, bool]:
    """Returns (character_or_None, matched) -- matched=False means the row
    should be skipped entirely (unrecognized name, e.g. a typo introduced
    while editing the file), never silently guessed into an existing
    character or created as a new one (character creation stays its own
    explicit flow, see project_character_management)."""
    name = (raw or "").strip()
    if name.lower() in unassigned_labels:
        return None, True
    for real in characters:
        if real.strip().lower() == name.lower():
            return real, True
    return None, False


def _rows_to_entries(header_cells: list[str], data_rows: list[list[str]], characters: list[str]) -> ParsedImport:
    columns: list[tuple[str, str] | None] = [_parse_header_cell(c) for c in header_cells]
    unassigned_labels = _all_unassigned_labels()
    result = ParsedImport()
    seen_unmatched: set[str] = set()

    for row in data_rows:
        if not row or not (row[0] or "").strip():
            continue
        char_raw = row[0]
        character, matched = _resolve_character(char_raw, characters, unassigned_labels)
        if not matched:
            if char_raw.strip() not in seen_unmatched:
                seen_unmatched.add(char_raw.strip())
                result.unmatched_characters.append(char_raw.strip())
            continue
        # +1 -- row[0] is the character name, row[1:] are the activity
        # cells that line up with columns[0:] one-for-one (real bug found
        # + fixed here: reading row[i] instead of row[i+1] silently shifted
        # every value one column to the right, e.g. a Daily cell's value
        # landing on the Weekly column and the real last column never
        # being read at all).
        for i, col in enumerate(columns):
            if col is None or i + 1 >= len(row):
                continue
            cell = (row[i + 1] or "").strip()
            if cell not in ("1", "0"):
                continue  # blank = "no task of this name for this character", not a real state
            schedule, title = col
            result.entries.append(ImportEntry(character=character, schedule=schedule, title=title, done=cell == "1"))

    return result


def parse_full_view_csv(text: str, characters: list[str]) -> ParsedImport:
    import csv
    import io

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return ParsedImport()
    header, data_rows = rows[0], rows[1:]
    return _rows_to_entries(header[1:], data_rows, characters)


def parse_full_view_xlsx(path, characters: list[str]) -> ParsedImport:
    """XLSX's header is TWO rows (row 1: merged type-group cells, row 2:
    per-activity titles) instead of CSV's single "[Type] Title" cell --
    forward-fills row 1's merged-cell value across the blank cells openpyxl
    reports for every column a merge spans except the first, then
    recombines into the same "[Type] Title" shape _parse_header_cell
    already handles, so both formats share one parsing path from there.

    One real asymmetry vs. CSV: build_full_view_xlsx marks "done" with a
    "✓" glyph but writes NOTHING for an open cell -- same blank as a
    genuinely not-applicable cell (only the fill color tells them apart
    visually, which openpyxl's values_only read can't recover). So a blank
    XLSX cell here is treated exactly like CSV's blank ("no task of this
    name for this character" -- leave alone), never as an explicit
    "revert to open". Only "✓" (done) is unambiguous in this format --
    CSV is the format to use for a genuine open<->done round trip in both
    directions."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    grid = list(ws.iter_rows(values_only=True))
    if len(grid) < 2:
        return ParsedImport()

    type_row, title_row = grid[0][1:], grid[1][1:]
    last_type = ""
    header_cells = []
    for type_cell, title_cell in zip(type_row, title_row):
        if type_cell not in (None, ""):
            last_type = str(type_cell).strip()
        title = "" if title_cell is None else str(title_cell).strip()
        header_cells.append(f"[{last_type}] {title}" if title else "")

    data_rows = []
    for raw_row in grid[2:]:
        if not raw_row or raw_row[0] in (None, ""):
            continue
        name = str(raw_row[0])
        cells = []
        for v in raw_row[1:]:
            if v in (None, ""):
                cells.append("")
            elif str(v).strip() == "✓":
                cells.append("1")
            else:
                cells.append("0")
        data_rows.append([name] + cells)

    return _rows_to_entries(header_cells, data_rows, characters)
