"""Builds the "Full View" HTML snapshot -- a bigger, browser-rendered
Charakter-x-Aktivität grid over the REAL current tasks/shopping state,
opened via QDesktopServices from MainWindow._on_full_view_requested()
(User-Wunsch, 2026-09-04, concrete follow-up to a browser mockup of the
Daily-Dashboard/Roster-Grid idea from GitHub issue #2 -- placed via its own
button in the ToDo screen's sort/filter row, "nimm bitte diesen Platz").

A snapshot, not a live view: ticking a box in the opened page doesn't write
back into the running app. Real two-way sync would need a small local HTTP
server -- intentionally not built yet, a read-only overview already covers
what was asked for (see the earlier design discussion this followed up on).

Visual language (colors/fonts/layout) is carried over from the approved
browser mockup almost verbatim, so this reads as the same feature rather
than a second, differently-styled one -- only the data source changed:
real serialize_card() rows instead of fabricated example data, and cells
are binary done/open/not-applicable (the real completion model has no
fractional "3/5" progress the mockup's example data used, so this doesn't
pretend to have one).
"""

import html
import json

# Real gap found + fixed (User-reported, 2026-09-04: "translations fehlen
# hier") -- build_full_view_html() accepted a `language` argument from the
# very first version but never actually used it anywhere; every string in
# the exported page was hardcoded German regardless of the app's real
# language setting. Self-contained dict + helper (same pattern as
# ui/flow/widgets/delete_confirm_dialog.py's own local _T/_t) rather than
# adding ~25 export-page-only keys to the shared core/translations.py.
_T: dict[str, dict[str, str]] = {
    "de": {
        "eyebrow": "ToDo · Full View",
        "subtitle": "Momentaufnahme deiner echten Aufgaben — Häkchen hier schreiben NICHT in die App zurück, dafür bitte die App selbst nutzen.",
        "columns_btn": "🧩 Spalten",
        "hidden_suffix": "versteckt",
        "characters": "Charaktere",
        "active_suffix": "aktiv",
        "missed_tile_label": "Verpasst (Gestern)",
        "missed_banner_text": "{n} verpasste Aktivitäten von gestern — {names}",
        "missed_banner_hide": "Ausblenden",
        "all_chars": "Alle Chars",
        "unassigned": "Ohne Charakter",
        "all_types": "Alle Typen",
        "status_all": "Status: Alle",
        "status_open": "Offen",
        "status_done": "Erledigt",
        "prio_all": "Priorität: Alle",
        "prio_high": "Hoch",
        "prio_middle": "Mittel",
        "prio_low": "Niedrig",
        "search_placeholder": "Aktivität suchen…",
        "character_col": "Charakter",
        "score_col": "Score",
        "legend_done": "Erledigt",
        "legend_na": "Keine Aufgabe dieses Namens für diesen Charakter",
        "legend_high_prio": "Hohe Priorität",
        "note_b": "Nur Ansicht:",
        "note_text": "diese Seite ist ein Export — Änderungen hier fließen nicht in dein Profil zurück.",
        "picker_title": "Spalten wählen",
        "picker_desc": "Jede Kategorie zeigt standardmäßig 7.",
        "picker_reset": "Zurücksetzen (7/7/7)",
        "picker_apply": "Übernehmen",
        "picker_visible_suffix": "sichtbar",
        "more_chip": "mehr",
        "cell_na_title": "Keine Aufgabe dieses Namens für {name}",
        "cell_title_done": "erledigt",
        "cell_title_open": "offen",
        "high_prio_title": "Priorität: Hoch",
    },
    "ru": {
        "eyebrow": "ToDo · Полный вид",
        "subtitle": "Снимок твоих реальных задач — отметки здесь НЕ сохраняются в приложение, для этого используй само приложение.",
        "columns_btn": "🧩 Столбцы",
        "hidden_suffix": "скрыто",
        "characters": "Персонажи",
        "active_suffix": "активно",
        "missed_tile_label": "Пропущено (вчера)",
        "missed_banner_text": "{n} пропущенных активностей за вчера — {names}",
        "missed_banner_hide": "Скрыть",
        "all_chars": "Все персонажи",
        "unassigned": "Без персонажа",
        "all_types": "Все типы",
        "status_all": "Статус: Все",
        "status_open": "Открыто",
        "status_done": "Выполнено",
        "prio_all": "Приоритет: Все",
        "prio_high": "Высокий",
        "prio_middle": "Средний",
        "prio_low": "Низкий",
        "search_placeholder": "Поиск активности…",
        "character_col": "Персонаж",
        "score_col": "Счёт",
        "legend_done": "Выполнено",
        "legend_na": "Нет задачи с этим названием у этого персонажа",
        "legend_high_prio": "Высокий приоритет",
        "note_b": "Только просмотр:",
        "note_text": "эта страница — экспорт, изменения здесь не попадают обратно в твой профиль.",
        "picker_title": "Выбрать столбцы",
        "picker_desc": "Каждая категория по умолчанию показывает 7.",
        "picker_reset": "Сбросить (7/7/7)",
        "picker_apply": "Применить",
        "picker_visible_suffix": "видно",
        "more_chip": "ещё",
        "cell_na_title": "Нет задачи с этим названием у {name}",
        "cell_title_done": "выполнено",
        "cell_title_open": "открыто",
        "high_prio_title": "Приоритет: Высокий",
    },
    "en": {
        "eyebrow": "ToDo · Full View",
        "subtitle": "Snapshot of your real tasks — checking a box here does NOT write back into the app, use the app itself for that.",
        "columns_btn": "🧩 Columns",
        "hidden_suffix": "hidden",
        "characters": "Characters",
        "active_suffix": "active",
        "missed_tile_label": "Missed (Yesterday)",
        "missed_banner_text": "{n} missed activities from yesterday — {names}",
        "missed_banner_hide": "Hide",
        "all_chars": "All Characters",
        "unassigned": "Unassigned",
        "all_types": "All Types",
        "status_all": "Status: All",
        "status_open": "Open",
        "status_done": "Done",
        "prio_all": "Priority: All",
        "prio_high": "High",
        "prio_middle": "Middle",
        "prio_low": "Low",
        "search_placeholder": "Search activity…",
        "character_col": "Character",
        "score_col": "Score",
        "legend_done": "Done",
        "legend_na": "No task with this name for this character",
        "legend_high_prio": "High priority",
        "note_b": "View only:",
        "note_text": "this page is an export — changes here don't flow back into your profile.",
        "picker_title": "Choose Columns",
        "picker_desc": "Each category shows 7 by default.",
        "picker_reset": "Reset (7/7/7)",
        "picker_apply": "Apply",
        "picker_visible_suffix": "visible",
        "more_chip": "more",
        "cell_na_title": "No task with this name for {name}",
        "cell_title_done": "done",
        "cell_title_open": "open",
        "high_prio_title": "Priority: High",
    },
}


def _lt(key: str, lang: str, **kwargs) -> str:
    text = _T.get(lang, _T["en"]).get(key, _T["en"].get(key, key))
    return text.format(**kwargs) if kwargs else text


_STYLE = """
@import url('https://fonts.googleapis.com/css2?family=Cinzel:wght@600;700&family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@500;600&display=swap');
:root{
  --bg:#0f172a; --surface:#1b2536; --surface-2:#212d42; --panel-inferno:rgba(59,15,15,0.55);
  --border:rgba(249,115,22,0.28); --border-soft:rgba(148,163,184,0.16);
  --text:#e8edf5; --text-dim:#94a3b8; --text-faint:#5c6b84;
  --accent:#fb923c; --accent-strong:#f97316;
  --daily:#22d3ee; --weekly:#a78bfa; --season:#fb923c;
  --low:#4ade80; --mid:#fbbf24; --high:#f87171;
  --shadow: 0 12px 32px rgba(0,0,0,0.35);
}
*{box-sizing:border-box;}
body{
  margin:0;
  background:
    radial-gradient(1200px 500px at 15% -10%, rgba(249,115,22,0.10), transparent 60%),
    radial-gradient(900px 500px at 100% 0%, rgba(167,139,250,0.06), transparent 55%),
    var(--bg);
  color:var(--text); font-family:'Inter',system-ui,sans-serif; min-height:100vh;
}
.num{font-family:'JetBrains Mono','Consolas',monospace;font-variant-numeric:tabular-nums;}
.shell{max-width:1360px;margin:0 auto;padding:28px 32px 64px;}
.top{display:flex;align-items:flex-end;justify-content:space-between;gap:24px;flex-wrap:wrap;margin-bottom:18px;}
.eyebrow{font-size:11px;letter-spacing:.14em;text-transform:uppercase;color:var(--accent);font-weight:600;margin:0 0 6px;}
h1{font-family:'Cinzel',serif;font-weight:700;font-size:28px;margin:0;text-wrap:balance;color:#fbead9;}
.subtitle{color:var(--text-dim);font-size:13px;margin-top:6px;max-width:56ch;}
.btn{
  border:1px solid var(--border);background:var(--panel-inferno);color:var(--text);
  font:inherit;font-weight:600;font-size:12.5px;padding:9px 15px;border-radius:9px;
  cursor:pointer;display:inline-flex;align-items:center;gap:7px;white-space:nowrap;
  text-decoration:none;
}
.btn:hover{border-color:var(--accent);background:rgba(249,115,22,0.16);}
.tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:14px;margin-bottom:20px;}
.tile{background:var(--surface);border:1px solid var(--border-soft);border-radius:12px;padding:14px 16px;}
.tile .k{font-size:11px;color:var(--text-dim);text-transform:uppercase;letter-spacing:.08em;font-weight:600;}
.tile .v{font-size:24px;font-weight:700;margin-top:4px;}
.tile .v small{font-size:13px;color:var(--text-dim);font-weight:500;}
.tile .bar{height:5px;border-radius:3px;background:rgba(148,163,184,0.16);margin-top:9px;overflow:hidden;}
.tile .bar i{display:block;height:100%;background:linear-gradient(90deg,#22d3ee,#a855f7);}
.tile.missed .v{color:var(--high);}
.missed-banner{display:flex;align-items:center;gap:10px;background:rgba(248,113,113,0.08);border:1px solid rgba(248,113,113,0.35);border-radius:10px;padding:10px 14px;margin-bottom:14px;font-size:12.5px;color:var(--text-dim);}
.missed-banner[hidden]{display:none;}
.missed-banner .dot{width:8px;height:8px;border-radius:50%;background:var(--high);flex:none;}
.missed-banner b{color:var(--text);font-weight:700;}
.missed-banner .hide{margin-left:auto;background:none;border:none;color:var(--text-dim);font:inherit;font-weight:600;cursor:pointer;text-decoration:underline;white-space:nowrap;flex:none;}
.missed-banner .hide:hover{color:var(--text);}
.toolbar{
  display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:14px;
  background:var(--surface);border:1px solid var(--border-soft);border-radius:12px;padding:10px 12px;
}
.chipset{display:flex;gap:6px;flex-wrap:wrap;}
.chip{
  border:1px solid var(--border-soft);background:transparent;color:var(--text-dim);
  font:inherit;font-size:12px;font-weight:600;padding:6px 11px;border-radius:999px;cursor:pointer;
}
.chip:hover{border-color:var(--accent);color:var(--text);}
.chip.on{background:rgba(251,146,60,0.16);border-color:var(--accent);color:#fed7aa;}
.toolbar .sep{width:1px;align-self:stretch;background:var(--border-soft);margin:0 2px;}
.spacer{flex:1;}
.search{
  display:flex;align-items:center;gap:7px;background:var(--surface-2);border:1px solid var(--border-soft);
  border-radius:9px;padding:7px 11px;min-width:180px;
}
.search input{background:none;border:none;color:var(--text);font:inherit;font-size:12.5px;outline:none;width:100%;}
.search input::placeholder{color:var(--text-faint);}
.toolbar select{
  background:var(--surface-2);border:1px solid var(--border-soft);color:var(--text);
  font:inherit;font-size:12.5px;font-weight:600;padding:7px 10px;border-radius:9px;cursor:pointer;
}
.grid-wrap{
  border:1px solid var(--border-soft);border-radius:14px;overflow:auto;background:var(--surface);
  box-shadow:var(--shadow);max-height:76vh;
}
table{border-collapse:separate;border-spacing:0;width:100%;min-width:900px;}
thead th{position:sticky;top:0;z-index:3;background:#151f30;padding:0;border-bottom:1px solid var(--border-soft);}
.grp-row th{padding:9px 10px 6px;font-size:10.5px;letter-spacing:.09em;text-transform:uppercase;font-weight:700;text-align:center;border-left:1px solid rgba(255,255,255,0.04);}
.grp-row th.g-daily{color:var(--daily);} .grp-row th.g-weekly{color:var(--weekly);} .grp-row th.g-season{color:var(--season);}
.grp-row th.corner{background:#151f30;}
.col-row th{padding:8px 10px 12px;font-size:12px;font-weight:600;color:var(--text-dim);text-align:center;min-width:104px;border-left:1px solid rgba(255,255,255,0.04);}
.col-row th .req{display:block;font-size:10px;color:var(--text-faint);font-weight:500;margin-top:2px;}
.col-row th.corner{position:sticky;left:0;z-index:4;background:#151f30;text-align:left;padding:8px 14px;min-width:200px;}
tbody th{position:sticky;left:0;z-index:2;background:var(--surface);text-align:left;padding:10px 14px;border-right:1px solid var(--border-soft);border-bottom:1px solid rgba(255,255,255,0.05);min-width:200px;}
tbody tr:nth-child(even) th{background:#1e2839;}
tbody tr:hover th{background:#252a36;}
tbody tr:nth-child(even) td{background-color:rgba(255,255,255,0.015);}
tbody tr:hover td{background-color:rgba(251,146,60,0.045);}
.char{display:flex;align-items:center;gap:10px;}
.avatar{width:30px;height:30px;border-radius:9px;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:12px;flex:none;color:#1a0d02;}
.char b{font-size:13px;font-weight:600;}
td{padding:8px;text-align:center;border-left:1px solid rgba(255,255,255,0.04);border-bottom:1px solid rgba(255,255,255,0.05);vertical-align:middle;}
.cell{display:inline-flex;flex-direction:column;align-items:center;gap:3px;width:100%;padding:6px 4px;border-radius:9px;border:1px solid transparent;position:relative;}
.cell .box{width:19px;height:19px;border-radius:6px;border:1.6px solid #56637a;display:flex;align-items:center;justify-content:center;}
.cell .frac{font-size:11px;color:var(--text-dim);font-weight:600;}
.cell.done .box{background:var(--low);border-color:var(--low);}
.cell.done .box:after{content:"✓";font-size:11px;color:#062e12;font-weight:700;}
.cell.done .frac{color:var(--low);}
.cell.na{opacity:0.28;}
.cell .flag{position:absolute;top:2px;right:2px;width:7px;height:7px;border-radius:50%;background:var(--high);box-shadow:0 0 0 2px var(--surface);}
.score{font-size:12px;} .score b{font-size:14px;}
.score.hi b{color:var(--low);} .score.mid b{color:var(--mid);} .score.lo b{color:var(--high);}
tfoot td{padding:9px 8px;text-align:center;font-size:11px;color:var(--text-faint);border-top:1px solid var(--border-soft);}
tfoot th{position:sticky;left:0;background:#151f30;text-align:left;padding:9px 14px;font-size:11px;color:var(--text-faint);font-weight:600;border-top:1px solid var(--border-soft);}
.legend{display:flex;gap:18px;flex-wrap:wrap;margin-top:16px;font-size:12px;color:var(--text-dim);}
.legend span{display:inline-flex;align-items:center;gap:7px;}
.legend i{width:12px;height:12px;border-radius:4px;display:inline-block;border:1.6px solid #56637a;}
.legend i.done{background:var(--low);border-color:var(--low);}
.legend i.flag{border-radius:50%;background:var(--high);border:none;width:8px;height:8px;}
.legend i.na{opacity:0.28;background:#56637a;}
.note{margin-top:22px;padding:12px 16px;border-radius:10px;background:var(--surface);border:1px dashed var(--border-soft);color:var(--text-dim);font-size:12px;line-height:1.6;}
.note b{color:var(--text);}
.overlay{position:fixed;inset:0;background:rgba(6,10,18,0.55);display:flex;align-items:flex-start;justify-content:center;padding-top:9vh;z-index:20;}
.overlay[hidden]{display:none;}
.picker{width:420px;max-height:76vh;display:flex;flex-direction:column;background:#121b2c;border:1px solid var(--border);border-radius:14px;box-shadow:var(--shadow);overflow:hidden;}
.picker header{display:flex;align-items:center;justify-content:space-between;padding:14px 16px;border-bottom:1px solid var(--border-soft);}
.picker header h2{font-family:'Cinzel',serif;font-size:15px;margin:0;color:#fbead9;font-weight:700;}
.picker header p{margin:2px 0 0;font-size:11.5px;color:var(--text-dim);}
.picker .x{background:none;border:none;color:var(--text-dim);font-size:16px;cursor:pointer;padding:4px;}
.picker-body{overflow:auto;padding:6px 8px;}
.pgroup{padding:8px 8px 4px;}
.pgroup h3{font-size:10.5px;letter-spacing:.09em;text-transform:uppercase;font-weight:700;margin:4px 4px 6px;display:flex;align-items:center;justify-content:space-between;}
.pgroup h3 span{font-weight:600;color:var(--text-faint);letter-spacing:0;text-transform:none;font-size:10.5px;}
.pgroup.daily h3{color:var(--daily);} .pgroup.weekly h3{color:var(--weekly);} .pgroup.season h3{color:var(--season);}
.prow{display:flex;align-items:center;gap:10px;padding:7px 6px;border-radius:8px;cursor:pointer;font-size:13px;}
.prow:hover{background:rgba(255,255,255,0.04);}
.prow input{accent-color:var(--accent-strong);width:15px;height:15px;flex:none;}
.picker footer{display:flex;justify-content:space-between;gap:8px;padding:10px 14px;border-top:1px solid var(--border-soft);}
"""

_SCRIPT = """
const TYPE_LABEL = {daily:'Daily', weekly:'Weekly', season:'Season'};
const HIGH_PRIO = new Set(COLS.filter(c=>c.hi).map(c=>c.key));

function renderHead(visibleCols){
  const thead = document.getElementById('gridHead');
  thead.innerHTML = '';
  const grpRow = document.createElement('tr');
  grpRow.className = 'grp-row';
  grpRow.appendChild(Object.assign(document.createElement('th'), {className:'corner'}));
  ['daily','weekly','season'].forEach(type=>{
    if(filters.type!=='all' && filters.type!==type) return;
    const shown = visibleCols.filter(c=>c.type===type);
    const hiddenCount = COLS.filter(c=>c.type===type && !c.v).length;
    const th = document.createElement('th');
    th.className = 'g-'+type;
    th.colSpan = shown.length || 1;
    th.innerHTML = TYPE_LABEL[type] + (hiddenCount ? `<br><span class="more-chip" data-open-type="${type}" style="display:inline-flex;align-items:center;gap:4px;margin-top:5px;border:1px dashed var(--border);background:rgba(255,255,255,0.03);color:var(--text-dim);font-size:10px;font-weight:700;padding:3px 8px;border-radius:999px;cursor:pointer;">+${hiddenCount} ${T.more_chip}</span>` : '');
    grpRow.appendChild(th);
  });
  grpRow.appendChild(Object.assign(document.createElement('th'), {className:'corner'}));
  thead.appendChild(grpRow);

  const colRow = document.createElement('tr');
  colRow.className = 'col-row';
  const corner = document.createElement('th');
  corner.className = 'corner'; corner.textContent = T.character_col;
  colRow.appendChild(corner);
  visibleCols.forEach(c=>{
    const th = document.createElement('th');
    if(HIGH_PRIO.has(c.key)) th.title = T.high_prio_title;
    th.innerHTML = `${c.label}<span class="req">${TYPE_LABEL[c.type]}</span>`;
    colRow.appendChild(th);
  });
  const scoreHead = document.createElement('th');
  scoreHead.className = 'corner'; scoreHead.textContent = T.score_col;
  colRow.appendChild(scoreHead);
  thead.appendChild(colRow);
}

function renderFoot(visibleCols){
  const tfoot = document.getElementById('gridFoot');
  const tr = document.createElement('tr');
  const th = document.createElement('th'); th.textContent = T.legend_done;
  tr.appendChild(th);
  visibleCols.forEach(c=>{
    let done=0, total=0;
    ROWS.forEach(row=>{const v=row.v[c.key]; if(v){total++; if(v.done) done++;}});
    const td = document.createElement('td'); td.className = 'num';
    td.textContent = total ? Math.round(100*done/total)+'%' : '–';
    tr.appendChild(td);
  });
  tr.appendChild(document.createElement('td'));
  tfoot.innerHTML = ''; tfoot.appendChild(tr);
}

let filters = {char:'all', type:'all', status:'all', prio:'all', q:''};

function render(){
  const visibleCols = COLS.filter(c=>c.v)
    .filter(c=>filters.type==='all'||c.type===filters.type)
    .filter(c=>filters.prio==='all'||c.prio===filters.prio)
    .filter(c=>!filters.q||c.label.toLowerCase().includes(filters.q));
  renderHead(visibleCols);
  renderFoot(visibleCols);
  const tbody = document.getElementById('gridBody');
  tbody.innerHTML = '';
  const rows = ROWS.filter(r => filters.char==='all' || r.id===filters.char);

  rows.forEach(row=>{
    const tr = document.createElement('tr');
    const initials = row.name.split(' ').map(w=>w[0]).slice(0,2).join('').toUpperCase() || '?';
    const th = document.createElement('th');
    th.innerHTML = `<div class="char"><div class="avatar" style="background:${row.color}">${initials}</div><div><b>${row.name}</b></div></div>`;
    tr.appendChild(th);

    let doneCount=0, totalCount=0, anyCell=false;
    visibleCols.forEach(c=>{
      const v = row.v[c.key];
      // Status filter acts on the CELL, not the column (User-Wunsch,
      // 2026-09-04: "diese beiden Features fehlen mir noch") -- "open"/
      // "done" only ever match a real (non-n/a) cell, matching the real
      // binary completion model (no fractional/"missed" states to filter
      // by here, unlike the earlier browser mockup's example data).
      if(filters.status!=='all'){
        if(!v) return;
        if(filters.status==='open' && v.done) return;
        if(filters.status==='done' && !v.done) return;
      }
      anyCell = true;
      const td = document.createElement('td');
      const flag = HIGH_PRIO.has(c.key) ? `<span class="flag" title="${T.legend_high_prio}"></span>` : '';
      if(!v){
        td.innerHTML = `<div class="cell na" title="${T.cell_na_title.replace('{name}', row.name)}"><span class="box"></span></div>`;
      } else {
        totalCount++; if(v.done) doneCount++;
        const st = v.done ? 'done' : '';
        const stTxt = v.done ? T.cell_title_done : T.cell_title_open;
        td.innerHTML = `<div class="cell ${st}" title="${row.name} · ${c.label}: ${stTxt}">${flag}<span class="box"></span></div>`;
      }
      tr.appendChild(td);
    });
    if(!anyCell && filters.status!=='all') return;

    const s = totalCount ? Math.round(100*doneCount/totalCount) : null;
    const cls = s===null ? '' : s>=80?'hi':s>=40?'mid':'lo';
    const tdScore = document.createElement('td');
    tdScore.innerHTML = s===null ? '–' : `<span class="score ${cls}"><b class="num">${s}%</b></span>`;
    tr.appendChild(tdScore);
    tbody.appendChild(tr);
  });

  const hiddenTotal = COLS.filter(c=>!c.v).length;
  document.getElementById('colBtnCount').textContent = hiddenTotal ? `(${hiddenTotal} versteckt)` : '';
  document.querySelectorAll('[data-open-type]').forEach(chip=>{
    chip.addEventListener('click', e=>{ e.stopPropagation(); openColPicker(chip.dataset.openType); });
  });
}

document.getElementById('charChips').addEventListener('click', e=>{
  const b = e.target.closest('.chip'); if(!b) return;
  [...b.parentElement.children].forEach(c=>c.classList.remove('on'));
  b.classList.add('on'); filters.char = b.dataset.v; render();
});
document.getElementById('typeChips').addEventListener('click', e=>{
  const b = e.target.closest('.chip'); if(!b) return;
  [...b.parentElement.children].forEach(c=>c.classList.remove('on'));
  b.classList.add('on'); filters.type = b.dataset.v; render();
});
document.getElementById('statusFilter').addEventListener('change', e=>{filters.status=e.target.value; render();});
document.getElementById('prioFilter').addEventListener('change', e=>{filters.prio=e.target.value; render();});
document.getElementById('search').addEventListener('input', e=>{filters.q=e.target.value.toLowerCase(); render();});

const colOverlay = document.getElementById('colOverlay');
let pending = null;
function buildColBody(scrollToType){
  pending = Object.fromEntries(COLS.map(c=>[c.key, c.v]));
  const body = document.getElementById('colBody');
  body.innerHTML = '';
  ['daily','weekly','season'].forEach(type=>{
    const group = document.createElement('div');
    group.className = 'pgroup ' + type;
    const shownCount = COLS.filter(c=>c.type===type && pending[c.key]).length;
    group.innerHTML = `<h3>${TYPE_LABEL[type]} <span>${shownCount} ${T.picker_visible_suffix}</span></h3>`;
    COLS.filter(c=>c.type===type).forEach(c=>{
      const row = document.createElement('label');
      row.className = 'prow';
      row.innerHTML = `<input type="checkbox" ${pending[c.key]?'checked':''} data-key="${c.key}"> ${c.label}`;
      row.querySelector('input').addEventListener('change', e=>{
        pending[c.key] = e.target.checked;
        group.querySelector('h3 span').textContent = COLS.filter(x=>x.type===type && pending[x.key]).length + ' ' + T.picker_visible_suffix;
      });
      group.appendChild(row);
    });
    body.appendChild(group);
    if(type===scrollToType) requestAnimationFrame(()=>group.scrollIntoView({block:'start'}));
  });
}
function openColPicker(scrollToType){ buildColBody(scrollToType); colOverlay.hidden = false; }
document.getElementById('colBtn').addEventListener('click', ()=>openColPicker(null));
document.getElementById('colClose').addEventListener('click', ()=>colOverlay.hidden = true);
colOverlay.addEventListener('click', e=>{ if(e.target === colOverlay) colOverlay.hidden = true; });
document.getElementById('colApply').addEventListener('click', ()=>{
  COLS.forEach(c=>{ if(c.key in pending) c.v = pending[c.key]; });
  colOverlay.hidden = true; render();
});
document.getElementById('colReset').addEventListener('click', ()=>{
  ['daily','weekly','season'].forEach(type=>{
    COLS.filter(c=>c.type===type).forEach((c,i)=>{ pending[c.key] = i < 7; });
  });
  buildColBody(null);
});

// Only present in the DOM when there's something real to show (User-
// Wunsch, 2026-09-05) -- guarded since most opens of this page have
// nothing missed at all.
const missedHideBtn = document.getElementById('missedHideBtn');
if(missedHideBtn){
  missedHideBtn.addEventListener('click', ()=>{
    document.getElementById('missedBanner').hidden = true;
  });
}

render();
"""

_ROLE_COLORS = ["#22d3ee", "#a78bfa", "#4ade80", "#f87171", "#fb923c", "#60a5fa", "#f472b6", "#facc15"]


def _esc(s: str) -> str:
    return html.escape(str(s or ""), quote=True)


def _compute_grid(rows: list[dict], characters: list[str], language: str = "en") -> tuple[list[dict], list[dict]]:
    """Shared by every exporter (HTML/CSV/XLSX) so they all agree on the
    same columns/rows. rows: serialize_card() dicts (both 'task' and
    'shopping' type) for every currently-loaded task/shopping card.
    characters: MainWindow.characters (real Flow-Map-node-backed names,
    see project_character_management memory). Groups same-titled cards
    into one column per (title, schedule) pair -- the closest real-data
    equivalent of the mockup's "one activity per column" idea, since
    recurring per-character tasks in this app share the same title/
    schedule by convention (e.g. a "Daily Commands" task exists once per
    character, all under that title)."""
    cols_by_key: dict[str, dict] = {}
    row_ids = list(characters) + ["__none__"]
    row_names = {c: c for c in characters}
    row_names["__none__"] = _lt("unassigned", language)
    matrix: dict[str, dict[str, dict]] = {rid: {} for rid in row_ids}
    # Highest priority wins if the same-titled task somehow carries
    # different priorities across characters (rare, but a column needs one
    # representative value for the priority filter/flag).
    prio_rank = {"low": 0, "middle": 1, "high": 2}
    best_prio: dict[str, str] = {}

    for card in rows:
        title = (card.get("title") or "").strip()
        if not title:
            continue
        schedule = card.get("schedule") or "daily"
        key = f"{schedule}::{title}"
        cols_by_key.setdefault(key, {"key": key, "type": schedule, "label": title, "hi": False, "prio": "low"})
        prio = card.get("priority") or "low"
        if prio_rank.get(prio, 0) >= prio_rank.get(best_prio.get(key, "low"), 0):
            best_prio[key] = prio
        char = card.get("character") or "__none__"
        if char not in matrix:
            char = "__none__"
        entry = matrix[char].get(key)
        done = bool(card.get("completed"))
        if entry is None or (done and not entry["done"]):
            matrix[char][key] = {"done": done}

    for key, prio in best_prio.items():
        cols_by_key[key]["prio"] = prio
        cols_by_key[key]["hi"] = prio == "high"

    cols = sorted(cols_by_key.values(), key=lambda c: (["daily", "weekly", "season"].index(c["type"]) if c["type"] in ("daily", "weekly", "season") else 3, c["label"]))
    by_type_seen: dict[str, int] = {}
    for c in cols:
        idx = by_type_seen.get(c["type"], 0)
        c["v"] = idx < 7
        by_type_seen[c["type"]] = idx + 1

    row_objs = []
    for i, rid in enumerate(row_ids):
        if rid == "__none__" and not matrix[rid]:
            continue
        row_objs.append({
            "id": rid, "name": row_names[rid],
            "color": _ROLE_COLORS[i % len(_ROLE_COLORS)],
            "v": matrix[rid],
        })
    return cols, row_objs


def _type_progress(cols: list[dict], row_objs: list[dict], type_: str) -> tuple[int, int]:
    """(done, total) among real, applicable cells of one schedule type --
    used for both the summary tiles and each column-group's own average,
    real numbers only (no fabricated "missed yesterday" style stat that
    would need historical tracking this app doesn't have yet)."""
    keys = [c["key"] for c in cols if c["type"] == type_]
    done = total = 0
    for row in row_objs:
        for key in keys:
            v = row["v"].get(key)
            if v is not None:
                total += 1
                if v["done"]:
                    done += 1
    return done, total


def build_full_view_html(
    rows: list[dict], characters: list[str], language: str = "en", basename: str = "roster_grid",
    missed_daily: list[dict] | None = None,
) -> str:
    """basename: shared filename stem for the CSV/Excel sibling files this
    page links to (User-reported, 2026-09-04: "wird immer noch nicht
    angezeigt" -- a fixed "roster_grid.html" name meant some browsers
    reused/focused an already-open tab from a PREVIOUS export instead of
    loading the freshly-written one. MainWindow._on_full_view_requested
    now writes every export under a timestamped name instead, which this
    parameter threads through so the CSV/Excel links stay in sync with
    whichever HTML file is actually being served this time).

    missed_daily: MainWindow.missed_daily_activities -- real cards still
    incomplete right before the LAST daily reset (User-Wunsch, 2026-09-05:
    "die verpassten Missionen fehlen noch in dem Roster Grid"). Real
    tracked data, not the earlier browser mockup's fabricated example
    stat -- see MainWindow._record_missed_daily_activities. None/empty
    just means nothing was missed (or no daily reset has happened yet),
    same as any other real-zero state elsewhere on this page."""
    cols, row_objs = _compute_grid(rows, characters, language)
    missed_daily = missed_daily or []

    char_chips = "".join(
        f'<button class="chip" data-v="{_esc(r["id"])}">{_esc(r["name"])}</button>' for r in row_objs
    )

    def tile(label: str, done: int, total: int) -> str:
        pct = round(100 * done / total) if total else 0
        return f"""<div class="tile">
      <div class="k">{_esc(label)}</div>
      <div class="v">{done}<small>/{total}</small></div>
      <div class="bar"><i style="width:{pct}%"></i></div>
    </div>"""

    daily_done, daily_total = _type_progress(cols, row_objs, "daily")
    weekly_done, weekly_total = _type_progress(cols, row_objs, "weekly")
    season_done, season_total = _type_progress(cols, row_objs, "season")
    char_count = len(characters)
    missed_tile_html = ""
    missed_banner_html = ""
    if missed_daily:
        missed_tile_html = f"""<div class="tile missed"><div class="k">{_esc(_lt("missed_tile_label", language))}</div><div class="v">{len(missed_daily)}</div></div>"""
        names = ", ".join(
            f"{_esc(m.get('title', ''))} ({_esc(m['character'])})" if m.get("character") else _esc(m.get("title", ""))
            for m in missed_daily
        )
        banner_text = _lt("missed_banner_text", language, n=len(missed_daily), names=names)
        missed_banner_html = f"""<div class="missed-banner" id="missedBanner">
    <span class="dot"></span><span>{banner_text}</span>
    <button class="hide" id="missedHideBtn">{_esc(_lt("missed_banner_hide", language))}</button>
  </div>"""
    tiles_html = f"""<div class="tiles">
    {tile("Daily", daily_done, daily_total)}
    {tile("Weekly", weekly_done, weekly_total)}
    {tile("Season", season_done, season_total)}
    <div class="tile"><div class="k">{_esc(_lt("characters", language))}</div><div class="v">{char_count}<small> {_esc(_lt("active_suffix", language))}</small></div></div>
    {missed_tile_html}
  </div>"""

    body = f"""<!doctype html><html><head><meta charset="utf-8">
<title>Roster Grid</title><style>{_STYLE}</style></head><body>
<div class="shell">
  <div class="top">
    <div>
      <p class="eyebrow">{_esc(_lt("eyebrow", language))}</p>
      <h1>Roster Grid</h1>
      <p class="subtitle">{_esc(_lt("subtitle", language))}</p>
    </div>
    <button class="btn" id="colBtn">{_esc(_lt("columns_btn", language))} <span id="colBtnCount"></span></button>
    <a class="btn" href="{basename}.csv">⬇ CSV</a>
    <a class="btn" href="{basename}.xlsx">⬇ Excel</a>
  </div>
  {tiles_html}
  {missed_banner_html}
  <div class="toolbar">
    <div class="chipset" id="charChips">
      <button class="chip on" data-v="all">{_esc(_lt("all_chars", language))}</button>
      {char_chips}
    </div>
    <div class="sep"></div>
    <div class="chipset" id="typeChips">
      <button class="chip on" data-v="all">{_esc(_lt("all_types", language))}</button>
      <button class="chip" data-v="daily">Daily</button>
      <button class="chip" data-v="weekly">Weekly</button>
      <button class="chip" data-v="season">Season</button>
    </div>
    <div class="spacer"></div>
    <select id="statusFilter">
      <option value="all">{_esc(_lt("status_all", language))}</option>
      <option value="open">{_esc(_lt("status_open", language))}</option>
      <option value="done">{_esc(_lt("status_done", language))}</option>
    </select>
    <select id="prioFilter">
      <option value="all">{_esc(_lt("prio_all", language))}</option>
      <option value="high">{_esc(_lt("prio_high", language))}</option>
      <option value="middle">{_esc(_lt("prio_middle", language))}</option>
      <option value="low">{_esc(_lt("prio_low", language))}</option>
    </select>
    <div class="search">🔍<input id="search" placeholder="{_esc(_lt("search_placeholder", language))}"></div>
  </div>
  <div class="grid-wrap">
    <table>
      <thead id="gridHead"></thead>
      <tbody id="gridBody"></tbody>
      <tfoot id="gridFoot"></tfoot>
    </table>
  </div>
  <div class="legend">
    <span><i class="done"></i>{_esc(_lt("legend_done", language))}</span>
    <span><i class="na"></i>{_esc(_lt("legend_na", language))}</span>
    <span><i class="flag"></i>{_esc(_lt("legend_high_prio", language))}</span>
  </div>
  <div class="note"><b>{_esc(_lt("note_b", language))}</b> {_esc(_lt("note_text", language))}</div>
</div>
<div class="overlay" id="colOverlay" hidden>
  <div class="picker">
    <header>
      <div><h2>{_esc(_lt("picker_title", language))}</h2><p>{_esc(_lt("picker_desc", language))}</p></div>
      <button class="x" id="colClose">✕</button>
    </header>
    <div class="picker-body" id="colBody"></div>
    <footer>
      <button class="btn" id="colReset">{_esc(_lt("picker_reset", language))}</button>
      <button class="btn" id="colApply">{_esc(_lt("picker_apply", language))}</button>
    </footer>
  </div>
</div>
<script>
const T = {json.dumps(_T.get(language, _T["en"]), ensure_ascii=False)};
const COLS = {json.dumps(cols, ensure_ascii=False)};
const ROWS = {json.dumps(row_objs, ensure_ascii=False)};
{_SCRIPT}
</script>
</body></html>"""
    return body


_TYPE_LABEL = {"daily": "Daily", "weekly": "Weekly", "season": "Season"}


def build_full_view_csv(rows: list[dict], characters: list[str], language: str = "en") -> str:
    """Same Charakter x Aktivität grid as build_full_view_html, as plain
    CSV -- opens directly in Excel/Sheets (User-Wunsch, 2026-09-04:
    "vielleicht eine Option, aus dieser Ansicht als Excel-Tabelle
    exportieren"). Unlike the browser view, this always includes EVERY
    column, not just the default-visible 7 per category -- the "7 per
    type" cap is a viewing convenience, not a real data limit, and a data
    export should hand over everything. "1"/"0"/"" per cell (done/open/not
    applicable) rather than a checkmark glyph, so a spreadsheet can sum a
    column directly."""
    import csv
    import io

    cols, row_objs = _compute_grid(rows, characters, language)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([_lt("character_col", language)] + [f"[{_TYPE_LABEL.get(c['type'], c['type'])}] {c['label']}" for c in cols])
    for row in row_objs:
        line = [row["name"]]
        for c in cols:
            v = row["v"].get(c["key"])
            line.append("" if v is None else ("1" if v["done"] else "0"))
        writer.writerow(line)
    return buf.getvalue()


def build_full_view_xlsx(rows: list[dict], characters: list[str], path, language: str = "en"):
    """Same grid as build_full_view_csv, written as a real formatted
    .xlsx via openpyxl (the "echte Formatierung" upside over CSV the user
    was offered, 2026-09-04: "Können wir dem User die Wahl zwischen XLSX
    und CSV geben?", then "kannst du das stylischer gestalten?" for a
    second pass) -- a merged type-group header row (colored to match the
    app's own Daily/Weekly/Season badge colors, white bold text) above a
    dark per-activity header row, green-filled done cells, soft-grey
    not-applicable cells, thin borders and light row banding throughout,
    gridlines hidden (real borders replace them), an AutoFilter on the
    header row, frozen header rows + first column, and column widths
    sized to their actual content instead of one fixed guess."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    cols, row_objs = _compute_grid(rows, characters, language)

    wb = Workbook()
    ws = wb.active
    ws.title = "Roster Grid"
    ws.sheet_view.showGridLines = False

    type_fill = {
        "daily": PatternFill("solid", fgColor="22D3EE"),
        "weekly": PatternFill("solid", fgColor="A78BFA"),
        "season": PatternFill("solid", fgColor="FB923C"),
    }
    header_fill = PatternFill("solid", fgColor="1E293B")
    band_fill = PatternFill("solid", fgColor="F1F5F9")
    done_fill = PatternFill("solid", fgColor="BBF7D0")
    na_fill = PatternFill("solid", fgColor="E2E8F0")
    corner_fill = PatternFill("solid", fgColor="0F172A")

    group_font = Font(bold=True, color="0F172A", size=11)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    corner_font = Font(bold=True, color="FFFFFF", size=12)
    name_font = Font(bold=True, color="0F172A")
    done_font = Font(bold=True, color="15803D")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", indent=1)
    thin = Side(style="thin", color="CBD5E1")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_cols = len(cols)
    last_col_letter = get_column_letter(n_cols + 1)

    # Corner cell -- one real merge across both header rows instead of a
    # blank row 2 underneath it, so "Charakter" sits centered in its own
    # 2-row-tall block like the browser grid's own sticky corner.
    ws.merge_cells("A1:A2")
    corner = ws.cell(row=1, column=1, value=_lt("character_col", language))
    corner.font = corner_font
    corner.alignment = center
    corner.fill = corner_fill
    ws.cell(row=2, column=1).fill = corner_fill

    col_idx = 2
    run_start = col_idx
    run_type = cols[0]["type"] if cols else None
    for c in cols + [{"type": None}]:  # sentinel to flush the last run
        if c["type"] != run_type:
            if run_type in type_fill:
                ws.merge_cells(start_row=1, start_column=run_start, end_row=1, end_column=col_idx - 1)
            cell = ws.cell(row=1, column=run_start, value=_TYPE_LABEL.get(run_type, run_type or ""))
            cell.font = group_font
            cell.alignment = center
            if run_type in type_fill:
                cell.fill = type_fill[run_type]
            for cc in range(run_start, col_idx):
                ws.cell(row=1, column=cc).border = grid_border
            run_start = col_idx
            run_type = c["type"]
        col_idx += 1

    for i, c in enumerate(cols, start=2):
        cell = ws.cell(row=2, column=i, value=c["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = grid_border

    for r, row in enumerate(row_objs, start=3):
        banded = (r % 2) == 0
        name_cell = ws.cell(row=r, column=1, value=row["name"])
        name_cell.font = name_font
        name_cell.alignment = left
        name_cell.fill = band_fill if banded else PatternFill("solid", fgColor="FFFFFF")
        name_cell.border = grid_border
        for i, c in enumerate(cols, start=2):
            v = row["v"].get(c["key"])
            cell = ws.cell(row=r, column=i)
            cell.alignment = center
            cell.border = grid_border
            if v is None:
                cell.fill = na_fill
            elif v["done"]:
                cell.value = "✓"
                cell.font = done_font
                cell.fill = done_fill
            else:
                cell.value = ""
                cell.fill = band_fill if banded else PatternFill("solid", fgColor="FFFFFF")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = max(22, max((len(r["name"]) for r in row_objs), default=0) + 6)
    for i, c in enumerate(cols, start=2):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(c["label"]) * 0.95 + 4)

    ws.auto_filter.ref = f"A2:{last_col_letter}{len(row_objs) + 2}"

    wb.save(path)
